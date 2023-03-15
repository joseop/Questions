import openpyxl
import random
"""
Se importan las librerias requeridas para leer archivos
se crean variables de los archivos xlsx
se crean variables de las hojas de dichos archivos
"""
preguntas_excel = openpyxl.load_workbook('assets/preguntas.xlsx')
puntajes_excel = openpyxl.load_workbook('assets/puntajes.xlsx')
hoja_preguntas = preguntas_excel.active
hoja_puntajes = puntajes_excel.active

"""
Este metodo genera la pregunta accediendo a la base de datos que esta en los archivos xlsx
crea una lista con la pregunta y las respuestas ubicadas en la misma fila de la pregunta
al final le añade otra lista que retorna el metodo puntajes() para trabajar con un solo mensaje 
"""
def generar_pregunta():
    row = random.randint(1, hoja_preguntas.max_row)
    pregunta = hoja_preguntas['A'+str(row)].value
    a = hoja_preguntas['B'+str(row)].value
    b = hoja_preguntas['C'+str(row)].value
    c = hoja_preguntas['D'+str(row)].value
    d = hoja_preguntas['E'+str(row)].value
    correcta = hoja_preguntas['F'+str(row)].value

    pregunta_respuestas = [pregunta, a, b, c, d, correcta]
    return (pregunta_respuestas+puntajes())

"""
Este metodo sirve para retornar una lista de los puntajes, guardada en un segundo archivos 
xlsx definido anteriormente

"""
def puntajes():
    puntajes = []
    for i in range(1, 51):
        puntajes.append(hoja_puntajes['A'+str(i)].value)
        puntajes.append(str(hoja_puntajes['B'+str(i)].value))
    return puntajes

"""
Este metodo actualiza los puntajes en la base de datos o el archivo xlsx recibidos como parametros
Comparamos los puntajes y los vamos ordenando de mayor a menor
"""
def actualizar_puntaje(nuevo_puntaje):
    lista_puntaje = []
    for i in range(1, 51):
        lista_puntaje.append((hoja_puntajes['A'+str(i)].value, hoja_puntajes['B'+str(i)].value))

    if nuevo_puntaje[1] > lista_puntaje[-1][1]:
        lista_puntaje = sorted(lista_puntaje[:-1] + [nuevo_puntaje], key=lambda x: x[1], reverse=True)

    for i in range(1, 51):
        hoja_puntajes['A'+str(i)] = lista_puntaje[i-1][0]
        hoja_puntajes['B'+str(i)] = lista_puntaje[i-1][1]

    puntajes_excel.save('assets/puntajes.xlsx') #Guarda la base de datos actualizada


def close():  #Cerrar las bases de datos
    preguntas_excel.close()
    puntajes_excel.close()
